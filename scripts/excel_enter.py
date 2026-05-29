from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelDataInserter:
    """
    Класс для вставки данных в существующий Excel-файл.
    Колонка A (ДСЕ изделия) записывается без перекодировки.
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)

    @staticmethod
    def fix_encoding(text):
        """Исправляет cp1252->cp1251 перекодировку."""
        if not isinstance(text, str):
            return text
        try:
            return text.encode('cp1252', errors='ignore').decode('cp1251', errors='ignore')
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text

    @staticmethod
    def fix_data_encoding(data):
        """Рекурсивно исправляет кодировку во всех строках словаря."""
        if isinstance(data, dict):
            return {k: ExcelDataInserter.fix_data_encoding(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [ExcelDataInserter.fix_data_encoding(item) for item in data]
        elif isinstance(data, str):
            return ExcelDataInserter.fix_encoding(data)
        else:
            return data

    def insert_data(self, data, sheet_name, headers=None, fix_encoding=True):
        """
        Вставляет данные в новый лист.
        Колонка A берётся из ключа верхнего уровня словаря — без перекодировки.
        """
        if fix_encoding:
            data = self.fix_data_encoding(data)

        # Удаляем/создаём лист
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        ws = self.wb.create_sheet(title=sheet_name)

        # Определяем заголовки
        if headers is None:
            first_item = None
            for product_value in data.values():
                for detail_value in product_value.values():
                    first_item = detail_value
                    break
                if first_item:
                    break
            headers = list(first_item.keys()) if first_item else []

        # Заголовки
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Данные — КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: колонка A из product_key, остальные из detail_value
        row_idx = 2
        for product_key, product_value in data.items():
            for detail_key, detail_value in product_value.items():
                for col_idx, header in enumerate(headers, 1):
                    if col_idx == 1:  # Колонка A — из оригинального ключа (без перекодировки)
                        value = product_key
                    else:
                        value = detail_value.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

        # Автоширина столбцов
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        self.wb.save(self.file_path)
        return self.file_path

    def close(self):
        self.wb.close()